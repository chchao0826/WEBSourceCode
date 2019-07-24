import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, Border, Side,PatternFill
from openpyxl.styles import Font
from app.views.Scheduling import SchedulingSQL_ZL_PMCHDR


# 查找文件夹是否存在
def findPath(path):
    if os.path.exists(path) is False:
        os.mkdir(path)

# 表头建立
def excelTitle():
    titleVar = (
        '顺序',
        '超时',
        '客户名',
        '布车号',
        '物料编号',
        'LOT',
        '卡号',
        '色号',
        '投胚',
        '上工段',
        '现工段',
        '下工段',
        '生管交期',
        '业务交期',
        '耗时',
        '营业课别',
        '工卡备注',
    )
    return titleVar

def CreateExcel():
    print('---------------=======+++++')
    data1 = SchedulingSQL_ZL_PMCHDR('预定')
    data2 = SchedulingSQL_ZL_PMCHDR('水洗1')
    data3 = SchedulingSQL_ZL_PMCHDR('水洗2')
    wb = Workbook()
    sheet1 = wb.create_sheet("预定")
    sheet2 = wb.create_sheet("水洗1")
    sheet3 = wb.create_sheet("水洗2")
    sheet1.append(excelTitle())
    sheet2.append(excelTitle())
    sheet3.append(excelTitle())

    print(data1)
    print('=============')
    print(data2)
    print('+++++++++++++')
    print(data3)
    print('------+++++')
    highlight = NamedStyle(name="highlight")
    
    a = 2
    b = 1

    for i in data1:
        for key in i:
            sheet1.cell(row = a, column = b, value = i[key])
            b += 1        
        b = 1
        a += 1
        print(i)

    for i in data2:
        for key in i:
            sheet2.cell(row = a, column = b, value = i[key])
            b += 1        
        b = 1
        a += 1
        print(i)

    for i in data3:
        for key in i:
            sheet3.cell(row = a, column = b, value = i[key])
            b += 1        
        b = 1
        a += 1
        print(i)

    bIsFile = os.path.isfile("%s.xlsx" %'色样')
    if bIsFile:
        os.remove("%s.xlsx" %'色样')
    wb.save("%s.xlsx" %'色样')

if __name__ == "__main__":
    CreateExcel()